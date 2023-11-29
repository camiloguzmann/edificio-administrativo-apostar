from typing import Any
import cv2,pytesseract , re
from django.db.models.query import QuerySet
import numpy as np
from django.shortcuts import render, redirect
from django.http import JsonResponse , HttpResponseRedirect , HttpResponse
from .forms import RegistroVisitanteForm , EmpleadoForm , UsuarioForm
from django.contrib import messages
from django.views.generic import ListView, TemplateView , CreateView , UpdateView
from .models import Empleado, Salida , Visitantes , Usuario
from django.shortcuts import get_object_or_404
from django.shortcuts import render, get_object_or_404
from django.urls import reverse, reverse_lazy
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime 
from django.utils import timezone
from openpyxl.utils import get_column_letter
from django.contrib.auth.views import PasswordResetView
from django.contrib.auth.mixins import LoginRequiredMixin , PermissionRequiredMixin
from django.views import View
from edificio_app.mixins import LoginYSuperStaffMixin
from django.contrib.auth.models import Permission


# Create your views here.

class indexView(LoginRequiredMixin,TemplateView):
    template_name = 'index.html'
    def get_context_data(self, **kwargs):
        context=super().get_context_data(**kwargs)
        context['titulo']='SISTEMA DE SEGURIDAD'
        return context

class CreateVisitanteFormView(LoginRequiredMixin,TemplateView):
    template_name = 'crearVisitante.html'
    form_class = RegistroVisitanteForm

    def get(self, request, *args, **kwargs):
        visitante_form = self.form_class()
        return render(request, self.template_name, {'form': visitante_form})

    def post(self, request, *args, **kwargs):
        visitante_form = self.form_class(request.POST)

        if visitante_form.is_valid():
            visitante_form.save()
            response_data = {
                'success': True,
                'redirect_url': reverse('salida'),  # O la URL a la que deseas redirigir
                'title': 'Éxito',
                'message': 'El Visitante se ha registrado exitosamente.',
                'icon': 'success',
            }
            return JsonResponse(response_data)

        return render(request, self.template_name, {'form': visitante_form})

def completarCedula(request):
    if request.method == 'POST':
        cedula = request.POST.get('cedula', None)
        # Procesar la imagen con OCR
        image = request.FILES['image'].read()
        nparr = np.frombuffer(image, np.uint8)
        imagen = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

        pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

        # Convertir a escala de grises
        gris = cv2.cvtColor(imagen, cv2.COLOR_BGR2GRAY)
        # Aplicar filtro de umbral
        umbral = cv2.adaptiveThreshold(gris, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 55, 25)
        # Configurar OCR
        config = '--psm 1'
        text = pytesseract.image_to_string(umbral, config=config)
        # Normalizar espacios en blanco
        text = ' '.join(text.split())

        # Patrón para buscar número de cédula
        cedula_pattern = re.compile(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}|\d{8})')
        cedula_match = cedula_pattern.search(text)
        cedula = cedula_match.group(1) if cedula_match else 'Número de cédula no encontrado'
        # extraer la cedula sin puntos
        cedula = ''.join(filter(str.isdigit, cedula))

        response_data = {'cedula': cedula}
        
        return JsonResponse(response_data)

class ObtenerDatosVisitanteView(LoginRequiredMixin, View):
    def get(self, request, cedula, *args, **kwargs):
        # Filtrar visitantes por la identificación proporcionada
        visitantes = Visitantes.objects.filter(identificacion=cedula)

        if visitantes.exists():
            # Tomar el primer visitante del conjunto de resultados
            primer_visitante = visitantes.first()

            # Devolver los datos del primer visitante en formato JSON
            data = {
                'existe': True,
                'nombres': primer_visitante.nombres,
                'apellidos': primer_visitante.apellidos,
                'celular': primer_visitante.celular,
            }
        else:
            # No hay visitantes con la identificación proporcionada
            data = {'existe': False}

        return JsonResponse(data)
    
class VisitantesSalidaView(LoginRequiredMixin,ListView):
    model = Visitantes
    template_name = 'salidaVisitantes.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'VISITANTES SALIDA'
        return context

    def get_queryset(self):
        # Devuelve solo los visitantes que no tienen una salida registrada
        return Visitantes.objects.exclude(salida__isnull=False)

class SalidaView(LoginRequiredMixin,ListView):
    template_name = 'salidaVisitantes.html'

    def post(self, request, visitante_id, *args, **kwargs):
        visitante = get_object_or_404(Visitantes, pk=visitante_id)

        # Verificar si ya existe una salida para el visitante
        if not Salida.objects.filter(visitante=visitante).exists():
            # Crear una instancia de Salida
            salida = Salida(visitante=visitante)
            salida.save()

            # Añadir un mensaje de éxito
            response_data = {
                'success': True,
                'title': '¿Estás seguro?',
                'text': 'Una vez que confirme, no podrá deshacer esta acción.',
                'icon': 'warning',
                'showCancelButton': True,
                'confirmButtonColor': '#d33',
                'cancelButtonColor': '#3085d6',
                'confirmButtonText': 'Dar salida',
                'cancelButtonText': 'Cancelar',
            }
        else:
            # Añadir un mensaje de advertencia si ya existe una salida
            response_data = {
                'success': False,
                'title': 'Advertencia',
                'text': 'El visitante ya ha registrado su salida previamente.',
                'icon': 'warning',
                'showCancelButton': False,
                'confirmButtonColor': '#3085d6',
                'confirmButtonText': 'OK',
            }

        # Devolver la respuesta como JSON
        return JsonResponse(response_data)

    def get_queryset(self):
        return Visitantes.objects.exclude(salida__isnull=False)

class EmpleadoslistView(LoginRequiredMixin,ListView):
    model = Empleado
    template_name = 'empleados.html'  
    context_object_name = 'empleados'
    paginate_by = 15

    def get_queryset(self):
        return Empleado.objects.all()
    
    def get_context_data(self, **kwargs):
        context=super().get_context_data(**kwargs)
        context['titulo']='EMPLEADOS'
        return context

class CreateEmpleadoFormView(LoginRequiredMixin,CreateView):
    template_name = 'empleados/createEmpleados.html'
    form_class = EmpleadoForm

    def get(self, request, *args, **kwargs):
        empleado_form = self.form_class()
        return render(request, self.template_name, {'form': empleado_form})

    def post(self, request, *args, **kwargs):
        empleado_form = self.form_class(request.POST)

        if empleado_form.is_valid():
            empleado_form.save()
            response_data = {
                'success': True,
                'redirect_url': reverse('empleados'),  # O la URL a la que deseas redirigir
                'title': 'Éxito',
                'message': 'El empleado se ha creado exitosamente.',
                'icon': 'success',
            }
            return JsonResponse(response_data)

        return render(request, self.template_name, {'form': empleado_form})

class EmpleadosEditView(LoginRequiredMixin,UpdateView):
    template_name = 'empleados/editarEmpleados.html'
    def get_context_data(self, **kwargs):
        context=super().get_context_data(**kwargs)
        context['titulo']='EDITAR EMPLEADOS'
        return context
   
class EditarEmpleadoView(LoginRequiredMixin, View):
    template_name = 'empleados/editarEmpleados.html'

    def get(self, request, empleado_id, *args, **kwargs):
        empleado = get_object_or_404(Empleado, id=empleado_id)
        form = EmpleadoForm(instance=empleado)
        return render(request, self.template_name, {'form': form, 'empleado': empleado})

    def post(self, request, empleado_id, *args, **kwargs):
        empleado = get_object_or_404(Empleado, id=empleado_id)
        form = EmpleadoForm(request.POST, instance=empleado)

        if form.is_valid():
            form.save()

            # Agregar un mensaje de éxito
            messages.success(request, 'Se ha editado el empleado correctamente.')

            # Configurar el response_data
            response_data = {
                'success': True,
                'message': 'Se ha editado el empleado correctamente.',
                'redirect_url': reverse('empleados'),  # O la URL a la que deseas redirigir
            }

            # Devolver la respuesta como JSON
            return JsonResponse(response_data)

        # Si el formulario no es válido, renderizar la página con errores
        return render(request, self.template_name, {'form': form, 'empleado': empleado})

class EliminarEmpleadoView(LoginRequiredMixin, View):
    def get(self, request, empleado_id, *args, **kwargs):
        empleado = get_object_or_404(Empleado, id=empleado_id)
        empleado.delete()
        messages.success(request, 'El empleado se eliminó correctamente.')

        # Devolver una respuesta JSON indicando el éxito de la eliminación
        return JsonResponse({'status': 'success'})

    def post(self, request, empleado_id, *args, **kwargs):
        # Manejar cualquier lógica adicional para el método POST si es necesario
        return redirect('empleados')
    
# def generar_excel(request):
#     tipo = request.POST.get('tipo',)
#     fecha_inicio_str = request.POST.get('EFI')
#     fecha_fin_str = request.POST.get('EFF')

#     # Crear un libro de trabajo y obtener la hoja de cálculo activa
#     libro_trabajo = Workbook()
#     hoja = libro_trabajo.active

#     fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d') if fecha_inicio_str else None
#     fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d') if fecha_fin_str else None

#     # Obtener la fecha actual en formato UTC
#     hoy = timezone.now().date()

#     # Crear encabezados en la hoja de cálculo
#     encabezados = ['Cedula_visitante', 'Nombre_visitante', 'Apellido_visitante', 'Celular_visitante', 'Empresa_visitante', 'Área_visitante', 'Empleado_visitante', 'Tipo de Equipo', 'Marca', 'Serial', 'Fecha entrada', 'Fecha salida']
#     hoja.append(encabezados)

#     for col_num, value in enumerate(encabezados, 1):
#         col_letter = get_column_letter(col_num)
#         hoja.column_dimensions[col_letter].width = max(len(str(value)) + 2, 10)  # Puedes ajustar el valor según tus necesidades


#     # Obtener datos según el tipo seleccionado
#     if tipo == '1':
#         # Filtrar por tipo de equipo y rango de fechas
#         queryset = Visitantes.objects.filter(
#             tipo_equipo__in=['Portatil', 'Tablet', 'Disco Duro'],
#             created_at__date__range=(fecha_inicio, fecha_fin)
#         )
#     elif tipo == '2':
#         # Filtrar por empleados sin equipo y rango de fechas
#         queryset = Visitantes.objects.filter(
#             tipo_equipo='',
#             created_at__date__range=(fecha_inicio, fecha_fin)
#         )
#     else:
#         # Filtrar por rango de fechas
#         queryset = Visitantes.objects.filter(
#             created_at__date__range=(fecha_inicio, fecha_fin)
#         )
    
#     # Filtrar empleados del día de hoy si no hay rango de fechas especificado
#     if not fecha_inicio and not fecha_fin:
#         queryset = queryset.filter(created_at__date=hoy)

#     # Agregar datos a la hoja de cálculo
#     for visitante in queryset:
#         created_at = visitante.created_at.replace(tzinfo=None) if visitante.created_at else None
#         fecha_salida = None
#         if hasattr(visitante, 'salida') and visitante.salida and visitante.salida.fecha_salida:
#             fecha_salida = visitante.salida.fecha_salida.replace(tzinfo=None)

#         fila = [
#             visitante.identificacion, visitante.nombres, visitante.apellidos, visitante.celular, visitante.empresa,
#             visitante.area_id.nombre, visitante.empleado_id.nombre, visitante.tipo_equipo, visitante.marca, visitante.serial,
#             created_at, fecha_salida
#         ]
#         hoja.append(fila)

#     # Establecer el estilo del encabezado en negrita
#     for celda in hoja[1]:
#         celda.font = Font(bold=True)

#     # Crear la respuesta HTTP con el archivo Excel adjunto
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=Ingresos.xlsx'
#     libro_trabajo.save(response)

#     return response

class ReportesView(LoginRequiredMixin,LoginYSuperStaffMixin,TemplateView):
    template_name = 'reportes.html'
    def get_context_data(self, **kwargs):
        context=super().get_context_data(**kwargs)
        context['titulo']='REPORTES'
        return context

class GenerarExcelView(LoginRequiredMixin,LoginYSuperStaffMixin,View):
    def post(self, request, *args, **kwargs):
        tipo = request.POST.get('tipo', '')
        fecha_inicio_str = request.POST.get('EFI', '')
        fecha_fin_str = request.POST.get('EFF', '')

        # Crear un libro de trabajo y obtener la hoja de cálculo activa
        libro_trabajo = Workbook()
        hoja = libro_trabajo.active

        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d') if fecha_inicio_str else None
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d') if fecha_fin_str else None

        # Obtener la fecha actual en formato UTC
        hoy = timezone.now().date()

        # Crear encabezados en la hoja de cálculo
        encabezados = ['Cedula_visitante', 'Nombre_visitante', 'Apellido_visitante', 'Celular_visitante', 'Empresa_visitante', 'Área_visitante', 'Empleado_visitante', 'Tipo de Equipo', 'Marca', 'Serial', 'Fecha entrada', 'Fecha salida']
        hoja.append(encabezados)

        for col_num, value in enumerate(encabezados, 1):
            col_letter = get_column_letter(col_num)
            hoja.column_dimensions[col_letter].width = max(len(str(value)) + 2, 10)  # Puedes ajustar el valor según tus necesidades

        # Obtener datos según el tipo seleccionado
        if tipo == '1':
            # Filtrar por tipo de equipo y rango de fechas
            queryset = Visitantes.objects.filter(
                tipo_equipo__in=['Portatil', 'Tablet', 'Disco Duro'],
                created_at__date__range=(fecha_inicio, fecha_fin)
            )
        elif tipo == '2':
            # Filtrar por empleados sin equipo y rango de fechas
            queryset = Visitantes.objects.filter(
                tipo_equipo='',
                created_at__date__range=(fecha_inicio, fecha_fin)
            )
        else:
            # Filtrar por rango de fechas
            queryset = Visitantes.objects.filter(
                created_at__date__range=(fecha_inicio, fecha_fin)
            )

        # Filtrar empleados del día de hoy si no hay rango de fechas especificado
        if not fecha_inicio and not fecha_fin:
            queryset = queryset.filter(created_at__date=hoy)

        # Agregar datos a la hoja de cálculo
        for visitante in queryset:
            created_at = visitante.created_at.replace(tzinfo=None) if visitante.created_at else None
            fecha_salida = None
            if hasattr(visitante, 'salida') and visitante.salida and visitante.salida.fecha_salida:
                fecha_salida = visitante.salida.fecha_salida.replace(tzinfo=None)

            fila = [
                visitante.identificacion, visitante.nombres, visitante.apellidos, visitante.celular, visitante.empresa,
                visitante.area_id.nombre, visitante.empleado_id.nombre, visitante.tipo_equipo, visitante.marca, visitante.serial,
                created_at, fecha_salida
            ]
            hoja.append(fila)

        # Establecer el estilo del encabezado en negrita
        for celda in hoja[1]:
            celda.font = Font(bold=True)

        # Crear la respuesta HTTP con el archivo Excel adjunto
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Ingresos.xlsx'
        libro_trabajo.save(response)

        return response
 
class UsersListView(LoginRequiredMixin,LoginYSuperStaffMixin,PermissionRequiredMixin,ListView):
    permission_required = ('edificio_app.view_usuario','edificio_app.add_usuario','edificio_app.change_usuario','edificio_app.delete_usuario')
    model = Usuario
    template_name = 'users.html'


    def get_queryset(self):
        # Filtra los usuarios que son staff
        return self.model.objects.filter(is_active=True)

    def get_context_data(self, **kwargs):
        context=super().get_context_data(**kwargs)
        context['titulo']='LISTA DE USUARIOS'
        return context

class UsersCreateView(LoginRequiredMixin, LoginYSuperStaffMixin, CreateView):
    model = Usuario
    template_name = 'users/createUsers.html'
    form_class = UsuarioForm
    success_url = reverse_lazy('users')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'CREAR USUARIOS'
        return context

    def form_valid(self, form):
        form.instance.set_password(form.cleaned_data['password1'])

        response = super().form_valid(form)
        usuario_creado = form.save()

        grupo_seleccionado = form.cleaned_data['tipo_usuario']
        usuario_creado.groups.add(grupo_seleccionado)

        if grupo_seleccionado.name == 'Administrador':
            usuario_creado.is_superuser = True
            usuario_creado.user_permissions.set(grupo_seleccionado.permissions.all())
            usuario_creado.save()

        messages.success(self.request, 'El usuario se ha creado exitosamente.')

        return JsonResponse({
            'title': 'Éxito',
            'message': 'El usuario se ha creado exitosamente.',
            'icon': 'success',
            'success': True,
            'redirect_url': self.success_url,
        })

    def form_invalid(self, form):
        messages.error(self.request, 'Hubo un error al procesar el formulario. Por favor, verifica los datos.')

        return JsonResponse({
            'title': 'Error',
            'message': 'Hubo un error al procesar el formulario. Por favor, verifica los datos.',
            'icon': 'error',
            'success': False,
        })

class EditarUsuarioView(LoginRequiredMixin, LoginYSuperStaffMixin,View):
    template_name = 'users/editarUsuarios.html'

    def get(self, request, usuario_id, *args, **kwargs):
        usuario = get_object_or_404(Usuario, id=usuario_id)
        form = UsuarioForm(instance=usuario)
        return render(request, self.template_name, {'form': form, 'usuario': usuario})

    def post(self, request, usuario_id, *args, **kwargs):
        usuario = get_object_or_404(Usuario, id=usuario_id)
        form = UsuarioForm(request.POST, instance=usuario)

        if form.is_valid():
            form.save(commit=False)
            grupo_seleccionado = form.cleaned_data['tipo_usuario']
            usuario.groups.set([grupo_seleccionado])

            if grupo_seleccionado.name == 'Administrador':
                usuario.is_staff = True
                admin_permissions = Permission.objects.filter(codename__startswith='admin')
                usuario.user_permissions.set(admin_permissions)
            else:
                usuario.is_staff = False
                usuario.user_permissions.clear()

            form.save()

            messages.success(request, 'Se ha editado el usuario correctamente.')
            response_data = {
                'success': True,
                'title': 'Éxito',
                'message': 'Se ha editado el usuario correctamente.',
                'icon': 'success',
                'redirect_url': reverse('users'),
            }
            return JsonResponse(response_data)

        response_data = {
            'success': False,
            'title': 'Error',
            'message': 'Ha ocurrido un error al editar el usuario.',
            'icon': 'error',
        }
        return JsonResponse(response_data)

class EliminarUsuarioView(LoginRequiredMixin,LoginYSuperStaffMixin,View):

    def get(self, request, usuario_id, *args, **kwargs):
        usuario = get_object_or_404(Usuario, id=usuario_id)
        usuario.delete()
        messages.success(request, 'El usuario se eliminó correctamente.')
        # Devolver una respuesta JSON indicando el éxito de la eliminación
        return JsonResponse({'status': 'success'})

    def post(self, request, usuario_id, *args, **kwargs):
        # Redirigir a la lista de usuarios en caso de una solicitud POST
        return redirect('users')

class CustomPasswordResetView(PasswordResetView):
    # Especifica el campo correcto en tu modelo de usuario personalizado
    email_field = 'email'
    extra_context = {'email_field': email_field}

